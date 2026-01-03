"""
Gerador de planilha de MDFs.
Processa PDFs, cruza com a planilha de escala e gera CSV/Excel alinhados ao cabecalho da BASE.csv.
"""

from __future__ import annotations

import glob
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Pastas onde os PDFs são procurados e arquivos padrão
PDF_SUBFOLDERS = ["SOROCABA", "ITU", "OUTRAS ORI-DES"]
BASE_FILENAME = "BASE.csv"
ESCALA_FALLBACK = "ESCALA MOTORISTAS 2025.xlsx"
CSV_PATTERN = "PLANILHA MDFS *.csv"
XLSX_PATTERN = "PLANILHA MDFS *.xlsx"
COMPACT_OUTPUT = True  # Quando True, reduz logs por item e evita scroll
VERBOSE = False         # Quando True, força logs detalhados mesmo em modo compacto
COLOR_OUTPUT = True     # Quando True, usa ANSI para colorir barras e títulos


# =============================================================================
# Utilidades e UI
# =============================================================================


def _supports_color() -> bool:
    return COLOR_OUTPUT and sys.stdout.isatty()


def _color(text: str, color: str, bold: bool = False) -> str:
    if not _supports_color():
        return text
    codes = {"cyan": "96", "green": "92", "yellow": "93", "magenta": "95"}
    code = codes.get(color)
    if not code:
        return text
    prefix = f"\033[1;{code}m" if bold else f"\033[{code}m"
    return f"{prefix}{text}\033[0m"


def print_banner() -> None:
    line = "=" * 60
    print("\n" + _color(line, "cyan", bold=True))
    print(_color("Gerador de Planilha de MDFs".center(60), "cyan", bold=True))
    print(_color("PDFs -> escala -> CSV/Excel".center(60), "green"))
    print(_color(line, "cyan", bold=True) + "\n")


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c)).upper()


def _print_progress(prefix: str, current: int, total: int, bar_len: int = 40) -> None:
    try:
        frac = float(current) / max(1, total)
        filled = int(round(bar_len * frac))
        color_on = _supports_color()
        green = "\033[92m" if color_on else ""
        cyan = "\033[96m" if color_on else ""
        reset = "\033[0m" if color_on else ""
        bar = f"{green}{'#' * filled}{reset}{'.' * (bar_len - filled)}"
        status = f"{cyan}{prefix}:{reset} |{bar}| {current}/{total}"
        sys.stdout.write(f"\r{status}")
        sys.stdout.flush()
        if current >= total:
            sys.stdout.write("\r" + " " * (len(status) + 2) + "\r")
    except Exception:
        pass


def _log_step(message: str) -> None:
    print(f"\n{_color('>>>', 'magenta', bold=True)} {message}")


def _log_detail(message: str) -> None:
    """Escreve logs de detalhe apenas se não estiver em modo compacto ou se verbose."""
    if VERBOSE or not COMPACT_OUTPUT:
        print(_color(message, "yellow"))


def _sheet_label(idx: Optional[int]) -> str:
    if idx == 1:
        return "DIA ATUAL"
    if idx == 2:
        return "DIA ANTERIOR"
    if idx:
        return f"ABA {idx}"
    return "?"


def solicitar_responsavel() -> str:
    """Obtém o nome do responsável via argumento ou terminal."""

    if len(sys.argv) > 1 and sys.argv[1].strip():
        resp = sys.argv[1].strip().upper()
        print(f"Responsavel definido (via argumento): {resp}")
        return resp

    while True:
        try:
            responsavel = input("Digite o nome do responsavel pela emissao: ").strip()
        except EOFError:
            responsavel = ""

        if not responsavel:
            print("ERRO: Campo obrigatorio. Digite o nome do responsavel")
            continue

        caracteres_validos = "áàâãéèêíïóôõöùûüçñÁÀÂÃÉÈÊÍÏÓÔÕÖÙÛÜÇÑ"
        tem_numero = any(c.isdigit() for c in responsavel)
        caracteres_ok = all(
            c.isalpha() or c.isspace() or c in caracteres_validos for c in responsavel
        )

        if tem_numero or not caracteres_ok:
            print("ERRO: Digite apenas letras (sem numeros ou caracteres especiais)")
            continue

        responsavel = responsavel.upper()
        print(f"Responsavel definido: {responsavel}")
        return responsavel


def encontrar_arquivo_escala(base_path: Path, fallback: str) -> Path:
    """Localiza qualquer arquivo que comece com ESCALA; se não existir, usa fallback."""
    for arquivo in base_path.iterdir():
        nome = arquivo.name.lower()
        if nome.startswith("escala") and nome.endswith(".xlsx"):
            return arquivo
    return base_path / fallback


def remover_arquivos_antigos(base_path: Path) -> None:
    """Remove planilhas geradas anteriormente na raiz para evitar confusão."""
    for padrao in (CSV_PATTERN, XLSX_PATTERN):
        for arquivo in glob.glob(str(base_path / padrao)):
            try:
                os.remove(arquivo)
                print(f"Arquivo antigo removido: {os.path.basename(arquivo)}")
            except Exception as exc:
                print(f"Erro ao remover {arquivo}: {exc}")


def calcular_data_arquivo() -> str:
    agora = datetime.now()
    fmt = "%d/%m/%Y"
    if agora.hour >= 22:
        return (agora + timedelta(days=1)).strftime(fmt)
    return agora.strftime(fmt)


def construir_paths(base_path: Path, data_label: str) -> Dict[str, Path]:
    file_label = data_label.replace("/", "-")
    base = {
        "csv_nome": f"PLANILHA MDFS {file_label}.csv",
        "xlsx_nome": f"PLANILHA MDFS {file_label}.xlsx",
    }
    base["csv_raiz"] = base_path / base["csv_nome"]
    base["xlsx_raiz"] = base_path / base["xlsx_nome"]
    base["csv_hist"] = base_path / "CSV" / base["csv_nome"]
    base["xlsx_hist"] = base_path / "EXCEL" / base["xlsx_nome"]
    return base


def _try_read_base_headers(path: Path) -> Tuple[List[str], Optional[str]]:
    """Lê apenas o cabeçalho do BASE.csv testando encodings comuns."""
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            df = pd.read_csv(path, encoding=enc, nrows=0)
            cols = [str(c).lstrip("\ufeff") for c in df.columns]
            print(f"[OK] BASE.csv com encoding: {enc}")
            return cols, enc
        except Exception:
            continue
    print("Erro: não foi possível ler BASE.csv")
    return [], None


def _read_sheet_rows(ws, stop_empty: int = 50) -> Tuple[List[str], List[Tuple]]:
    data_iter = ws.values
    headers = next(data_iter)
    headers = [str(h).strip() if h is not None else "" for h in headers]
    rows = []
    consecutive_empty = 0
    for row in data_iter:
        is_empty = all(c is None or str(c).strip() == "" for c in row)
        if is_empty:
            consecutive_empty += 1
            if consecutive_empty >= stop_empty:
                break
            continue
        consecutive_empty = 0
        rows.append(row)
    return headers, rows


def carregar_escala(excel_file: Path) -> Tuple[pd.DataFrame, List[int]]:
    """Carrega até duas abas visíveis (atual e anterior) e devolve linhas + índice da aba."""
    try:
        try:
            wb = load_workbook(excel_file, data_only=True, read_only=True)
        except Exception:
            print("[WARN] Leitura read_only falhou; usando modo normal...")
            wb = load_workbook(excel_file, data_only=True)
    except Exception as exc:
        print(f"Erro ao abrir a planilha de escala: {exc}")
        return pd.DataFrame(), []

    headers_main: Optional[List[str]] = None
    all_rows: List[Tuple] = []
    sheet_index: List[int] = []
    visible_idx = 0

    for ws in wb.worksheets[:2]:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        visible_idx += 1
        headers, rows = _read_sheet_rows(ws)
        if headers_main is None:
            headers_main = headers
        for row in rows:
            all_rows.append(row)
            sheet_index.append(visible_idx)
        print(f"  {_sheet_label(visible_idx)} ({ws.title}): linhas lidas {len(rows)}")

    if headers_main is None:
        headers_main = []
    print(f"Total de linhas carregadas: {len(all_rows)}")
    return pd.DataFrame(all_rows, columns=headers_main), sheet_index


@dataclass
class MotoristaInfo:
    nome: str
    nome_completo: str
    escala: str
    frota: str
    gpid: str
    cpf: str
    sheet_idx: Optional[int]


def preparar_motoristas(df: pd.DataFrame, row_sheet_index: List[int]) -> Tuple[Dict[str, MotoristaInfo], Dict[str, Tuple[str, Optional[int]]]]:
    """Monta o dicionário de motoristas e aliases normalizados a partir da planilha de escala."""
    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame()

    col_map = {_normalize(c): c for c in df.columns}

    def _get(row: pd.Series, desired: str) -> str:
        real = col_map.get(_normalize(desired))
        if real is None:
            return ""
        try:
            val = row.get(real, "")
            return "" if pd.isna(val) else str(val).strip()
        except Exception:
            return ""

    motoristas: Dict[str, MotoristaInfo] = {}
    aliases: Dict[str, Tuple[str, Optional[int]]] = {}

    for idx, (_, row) in enumerate(df.iterrows(), start=1):
        sheet_idx = row_sheet_index[idx - 1] if idx - 1 < len(row_sheet_index) else None
        _print_progress("Processando motoristas", idx, len(df))

        nome_bruto = _get(row, "MOTORISTA") or _get(row, "NOME")
        nome_limpo = re.sub(r"\s*\(.*?\)", "", nome_bruto or "").strip()
        if not nome_limpo or nome_limpo in motoristas:
            continue

        info = MotoristaInfo(
            nome=nome_limpo,
            nome_completo=_get(row, "NOME COMPLETO") or _get(row, "NOME"),
            escala=_get(row, "ESCALA"),
            frota=_get(row, "FROTA"),
            gpid=_get(row, "GPID"),
            cpf=_get(row, "CPF"),
            sheet_idx=sheet_idx,
        )
        motoristas[nome_limpo] = info

        alias_raw = re.sub(r"\s*\(.*?\)", "", _get(row, "NOME")).strip()
        if alias_raw:
            alias_key = _normalize(re.sub(r"\d+", "", alias_raw))
            nome_key = _normalize(re.sub(r"\d+", "", nome_limpo))
            if alias_key and alias_key != nome_key:
                aliases.setdefault(alias_key, (nome_limpo, sheet_idx))

    print(f"Motoristas encontrados: {len(motoristas)}")
    return motoristas, aliases


def listar_pdfs(base_folder: Path, subfolders: Iterable[str]) -> Tuple[List[str], Dict[str, Tuple[str, Path]]]:
    """Varre subpastas e registra localização de cada PDF pelo nome normalizado."""
    pdfs: List[str] = []
    localizacao: Dict[str, Tuple[str, Path]] = {}

    for sub in subfolders:
        pasta = base_folder / sub
        if not pasta.exists():
            print(f"  [{sub}] pasta não existe")
            continue
        arquivos = [f for f in os.listdir(pasta) if f.lower().endswith(".pdf")]
        if not arquivos:
            _log_detail(f"  [{sub}] nenhum arquivo")
            continue
        print(f"  [{sub}] {len(arquivos)} arquivo(s)")
        for arq in arquivos:
            nome = arq.replace(".pdf", "").replace(".PDF", "")
            pdfs.append(nome)
            localizacao[nome.upper()] = (sub, pasta / arq)
    print(f"\nTotal: {len(pdfs)} PDFs")
    return pdfs, localizacao


def extrair_dt_do_pdf(caminho_pdf: Path) -> str:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if texto:
                    match = re.search(r"DT:\s*[\"\']?(\d+)[\"\']?", texto, re.IGNORECASE)
                    if match:
                        return match.group(1).strip()
    except Exception as exc:
        print(f"  Erro ao ler PDF {caminho_pdf}: {exc}")
    return ""


def extrair_cte_do_pdf(caminho_pdf: Path) -> str:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if texto:
                    match = re.search(r"CTE:\s*[\"\']?(\d+)[\"\']?", texto, re.IGNORECASE)
                    if match:
                        return match.group(1).strip()
    except Exception as exc:
        print(f"  Erro ao ler PDF {caminho_pdf}: {exc}")
    return ""


def extrair_mdfe_do_pdf(caminho_pdf: Path) -> str:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if texto:
                    match = re.search(r"Modelo\s+Série\s+Número.*?\n.*?(\d{6})", texto, re.IGNORECASE | re.DOTALL)
                    if match:
                        return match.group(1).strip()
                    match = re.search(r"Número[:\s]+(\d{6})", texto, re.IGNORECASE)
                    if match:
                        return match.group(1).strip()
    except Exception as exc:
        print(f"  Erro ao ler PDF {caminho_pdf}: {exc}")
    return ""


def extrair_hora_mdfe_do_pdf(caminho_pdf: Path) -> str:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if texto:
                    match = re.search(r"\d{2}/\d{2}/\d{4}\s+(\d{2}:\d{2}:\d{2})", texto)
                    if match:
                        return match.group(1).strip()
    except Exception as exc:
        print(f"  Erro ao ler PDF {caminho_pdf}: {exc}")
    return ""


def extrair_carreta_cavalo_do_pdf(caminho_pdf: Path) -> Tuple[str, str]:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if not texto:
                    continue
                linhas = texto.split("\n")
                for i, linha in enumerate(linhas):
                    if "Placa" in linha and "RNTRC" in linha:
                        carreta = ""
                        cavalo = ""
                        if i + 1 < len(linhas):
                            primeira = linhas[i + 1].strip().split()
                            carreta = primeira[0] if primeira else ""
                        if i + 2 < len(linhas):
                            segunda = linhas[i + 2].strip().split()
                            cavalo = segunda[0] if segunda else ""
                        return carreta, cavalo
    except Exception as exc:
        print(f"  Erro ao ler PDF {caminho_pdf}: {exc}")
    return "", ""


def extrair_nf_do_pdf(caminho_pdf: Path) -> str:
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if texto:
                    match = re.search(r"NF:\s*(\d+(?:/\d+)*)", texto, re.IGNORECASE)
                    if match:
                        return match.group(1).strip()
    except Exception as exc:
        print(f"  Erro ao ler PDF {caminho_pdf}: {exc}")
    return ""


def processar_pdfs_extrair_dados(pdfs: List[str], pdf_localizacao: Dict[str, Tuple[str, Path]]) -> Dict[str, Dict[str, str]]:
    """Percorre PDFs e extrai campos específicos usando os extratores dedicados."""
    dados = {"dt": {}, "cte": {}, "mdfe": {}, "hora_mdfe": {}, "carreta": {}, "cavalo": {}, "nf": {}}
    _log_step("Extraindo dados dos PDFs...")
    total = len(pdfs)
    processed = 0
    for idx, pdf_nome in enumerate(pdfs, start=1):
        _print_progress("Extraindo PDFs", idx, total)
        pdf_limpo = re.sub(r"\s*\(.*?\)", "", pdf_nome).strip().upper()
        if pdf_limpo not in pdf_localizacao:
            continue
        subfolder, caminho_pdf = pdf_localizacao[pdf_limpo]
        dados["dt"][pdf_limpo] = extrair_dt_do_pdf(caminho_pdf)
        dados["cte"][pdf_limpo] = extrair_cte_do_pdf(caminho_pdf)
        dados["mdfe"][pdf_limpo] = extrair_mdfe_do_pdf(caminho_pdf)
        dados["hora_mdfe"][pdf_limpo] = extrair_hora_mdfe_do_pdf(caminho_pdf)
        carreta, cavalo = extrair_carreta_cavalo_do_pdf(caminho_pdf)
        dados["carreta"][pdf_limpo] = carreta
        dados["cavalo"][pdf_limpo] = cavalo
        dados["nf"][pdf_limpo] = extrair_nf_do_pdf(caminho_pdf)
        if any([dados["dt"].get(pdf_limpo), dados["cte"].get(pdf_limpo), dados["mdfe"].get(pdf_limpo)]):
            _log_detail(f"  {pdf_nome} [{subfolder}] processado")
            processed += 1
    print(f"Resumo da extração de PDFs: {processed}/{total} processados")
    return dados


def _match_motorista(
    pdf_norm: str,
    motoristas: Dict[str, MotoristaInfo],
    aliases: Dict[str, Tuple[str, Optional[int]]],
    prefer_sheet_one: bool,
) -> Tuple[Optional[MotoristaInfo], Optional[int]]:
    """Casa o nome do PDF com motorista da escala, priorizando aba 1 quando requerido (ITU)."""
    def _match_pool(pool: Iterable[Tuple[str, MotoristaInfo]], predicate) -> Tuple[Optional[MotoristaInfo], Optional[int]]:
        for nome, info in pool:
            if not predicate(info.sheet_idx):
                continue
            nome_norm = _normalize(nome)
            tokens = nome_norm.split()
            if pdf_norm == nome_norm or pdf_norm in tokens or nome_norm.startswith(pdf_norm) or pdf_norm in nome_norm:
                return info, info.sheet_idx
        return None, None

    pool_primary = list(motoristas.items())
    if prefer_sheet_one:
        info, idx = _match_pool(pool_primary, lambda s: s == 1)
        if info:
            return info, idx
    info, idx = _match_pool(pool_primary, lambda _: True)
    if info:
        return info, idx

    alias_candidates = []
    for alias_key, (nome_real, sheet_idx) in aliases.items():
        alias_candidates.append((alias_key, motoristas.get(nome_real), sheet_idx))
    for alias_key, info, sheet_idx in alias_candidates:
        if info is None:
            continue
        alias_norm = alias_key
        if alias_norm == pdf_norm or alias_norm.startswith(pdf_norm) or pdf_norm in alias_norm or pdf_norm in alias_norm.split():
            return info, sheet_idx
    return None, None


def iniciar_linha(colunas_base: List[str]) -> Dict[str, str]:
    return {col: "" for col in colunas_base}


def montar_registros(
    pdfs: List[str],
    pdf_localizacao: Dict[str, Tuple[str, Path]],
    motoristas: Dict[str, MotoristaInfo],
    aliases: Dict[str, Tuple[str, Optional[int]]],
    dados_pdfs: Dict[str, Dict[str, str]],
    colunas_base: List[str],
    data_label: str,
    responsavel: str,
) -> List[Dict[str, str]]:
    """Cria as linhas alinhadas ao BASE.csv aplicando as regras de negócio por origem."""
    registros: List[Dict[str, str]] = []
    matched = 0
    not_found = 0
    resumo_localizados: List[Tuple[str, Optional[int], str, str, str, str]] = []

    for pdf_nome in pdfs:
        pdf_limpo = re.sub(r"\s*\(.*?\)", "", pdf_nome).strip().upper()
        pdf_norm = _normalize(re.sub(r"\d+", "", pdf_limpo))

        subfolder = None
        if pdf_limpo in pdf_localizacao:
            subfolder, _ = pdf_localizacao[pdf_limpo]

        info_motorista, sheet_idx = _match_motorista(
            pdf_norm,
            motoristas,
            aliases,
            prefer_sheet_one=(subfolder == "ITU"),
        )

        if not info_motorista:
            print(f"X {pdf_limpo} não encontrado")
            not_found += 1
            continue

        linha = iniciar_linha(colunas_base)

        origem = destino = ""
        if subfolder in ("ITU", "SOROCABA"):
            origem, destino = subfolder, "DHL"

        linha["DATA"] = data_label
        linha["HORARIO (P2)"] = ""
        linha["DT"] = dados_pdfs["dt"].get(pdf_limpo, "")
        linha["STATUS (P2)"] = "FATURADO"
        linha["MOTORISTA"] = info_motorista.nome
        linha["NOME COMPLETO"] = info_motorista.nome_completo
        linha["GPID"] = info_motorista.gpid
        linha["CPF"] = info_motorista.cpf
        linha["HORA ESCALA (P2)"] = info_motorista.escala
        linha["HORA APRESENTACAO (P2)"] = ""
        linha["MOTIVO ATRASO (P2)"] = ""
        linha["CTE (P2)"] = dados_pdfs["cte"].get(pdf_limpo, "")
        linha["N MDFE (P2)"] = dados_pdfs["mdfe"].get(pdf_limpo, "")
        linha["HORA MDFE (P2)"] = dados_pdfs["hora_mdfe"].get(pdf_limpo, "")
        linha["EMITO POR (P2)"] = responsavel
        linha["ORIGEM (ESCALA)"] = origem
        linha["DESTINO (ESCALA)"] = destino
        linha["FROTA (P2)"] = info_motorista.frota
        linha["CAVALO (P2)"] = dados_pdfs["cavalo"].get(pdf_limpo, "")
        linha["CARRETA (P2)"] = dados_pdfs["carreta"].get(pdf_limpo, "")
        linha["NF (P2)"] = dados_pdfs["nf"].get(pdf_limpo, "")
        linha["ONE"] = ""
        linha["SAP"] = ""
        linha["RESPONSAVEL P2"] = responsavel
        linha["HORA ENTREGA INBOUND"] = ""
        linha["RESP. ENTREGA INBOUND"] = ""
        linha["OBSERVACOES (P2)"] = ""
        linha["DATA/ HORA SAIDA"] = ""
        linha["COD MOT"] = ""

        if subfolder == "SOROCABA":
            linha["MOTIVO ATRASO (P2)"] = "VETADO ANTECIPACAO DE MDF"
            linha["HORA ESCALA (P2)"] = ""

        registros.append(linha)
        matched += 1
        resumo_localizados.append(
            (
                info_motorista.nome,
                sheet_idx,
                info_motorista.frota or "-",
                subfolder or "-",
                dados_pdfs["dt"].get(pdf_limpo, "-") or "-",
                dados_pdfs["carreta"].get(pdf_limpo, "-") or "-",
            )
        )
        aba_info = _sheet_label(sheet_idx)
        if subfolder:
            _log_detail(f"[OK] {pdf_limpo} → {info_motorista.nome} ({aba_info}, pasta {subfolder})")
        else:
            _log_detail(f"[OK] {pdf_limpo} → {info_motorista.nome} ({aba_info})")

    print(_color(f"\nCorrespondência de motoristas: {matched} correspondências, {not_found} não localizados", "magenta", bold=True))
    if resumo_localizados:
        titulo = _color("Resumo de motoristas localizados", "green", bold=True)
        cabecalho = f"{'Nome':<26} | {'Dia':<12} | {'Local':<10} | {'Frota':<8} | {'DT':<10} | {'Carreta':<10}"
        print(titulo)
        print(_color(cabecalho, "cyan"))
        print("-" * len(cabecalho))

        limite = 40
        for nome, sheet, frota, local, dt_val, carreta_val in resumo_localizados[:limite]:
            dia = _sheet_label(sheet)
            linha = f"{nome:<26} | {dia:<12} | {local:<10} | {frota:<8} | {dt_val:<10} | {carreta_val:<10}"
            print(linha)

        if len(resumo_localizados) > limite:
            print(f"... (+{len(resumo_localizados) - limite} restantes)")
    return registros


def salvar_saidas(df: pd.DataFrame, colunas_base: List[str], paths: Dict[str, Path]) -> None:
    """Garante colunas na ordem do BASE e persiste CSV/Excel na raiz e histórico."""
    missing_cols = [c for c in colunas_base if c not in df.columns]
    for col in missing_cols:
        df[col] = ""
    df = df[colunas_base]

    try:
        df.to_csv(paths["csv_raiz"], index=False, encoding="latin-1")
        print(f"[OK] CSV criado: {paths['csv_nome']}")
    except Exception as exc:
        print(f"[ERRO] CSV raiz: {exc}")

    try:
        df.to_csv(paths["csv_hist"], index=False, encoding="latin-1")
        print(f"[OK] CSV arquivado: CSV/{paths['csv_nome']}")
    except Exception as exc:
        print(f"[ERRO] CSV histórico: {exc}")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        for col_num, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=col_name)
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        wb.save(paths["xlsx_raiz"])
        print(f"[OK] Excel criado: {paths['xlsx_nome']}")
        wb.save(paths["xlsx_hist"])
        print(f"[OK] Excel arquivado: EXCEL/{paths['xlsx_nome']}")
    except Exception as exc:
        print(f"[ERRO] Excel: {exc}")


def main() -> None:
    """Orquestra a execução end-to-end: entrada, leitura, casamento, geração de saída."""
    print_banner()
    responsavel = solicitar_responsavel()
    print(f"Responsavel: {responsavel}\n")
    print("Iniciando processamento...")
    print("=" * 60 + "\n")

    base_path = Path(__file__).parent
    pdf_base = base_path / "MDFs geradas"

    escala_path = encontrar_arquivo_escala(base_path, ESCALA_FALLBACK)
    print("Buscando arquivo de escala...")
    if escala_path.exists():
        print(f"Encontrado: {escala_path.name}\n")
    else:
        print(f"Aviso: arquivo de escala não encontrado, tentativa: {escala_path}\n")

    data_label = calcular_data_arquivo()
    print(f"Horário de execução: {datetime.now().strftime('%H:%M')}")
    print(f"Data usada no arquivo: {data_label}")

    paths = construir_paths(base_path, data_label)

    _log_step("Limpando arquivos antigos...")
    remover_arquivos_antigos(base_path)

    _log_step("Carregando dados base...")
    colunas_base, _ = _try_read_base_headers(base_path / BASE_FILENAME)
    if not colunas_base:
        print("Encerrando: BASE.csv não foi lida.")
        return

    _log_step("Carregando planilha de escala (aba atual e anterior visíveis)...")
    df_escala, row_sheet_index = carregar_escala(escala_path)

    _log_step("Processando motoristas...")
    motoristas, aliases = preparar_motoristas(df_escala, row_sheet_index)

    _log_step("Localizando PDFs...")
    pdfs, pdf_localizacao = listar_pdfs(pdf_base, PDF_SUBFOLDERS)

    dados_pdfs = processar_pdfs_extrair_dados(pdfs, pdf_localizacao)

    _log_step("Processando dados...")
    registros = montar_registros(
        pdfs,
        pdf_localizacao,
        motoristas,
        aliases,
        dados_pdfs,
        colunas_base,
        data_label,
        responsavel,
    )

    if not registros:
        print("\n" + "=" * 60)
        print("ERRO - Nenhum motorista foi encontrado!")
        print("=" * 60)
        print(
            "Nenhum motorista foi encontrado. Verifique:"
            "\n  - Se existe arquivo começando com 'ESCALA'"
            "\n  - Se existem PDFs nas subpastas"
            "\n  - Se os nomes dos PDFs correspondem aos motoristas"
        )
        return

    _log_step("Gerando arquivos...")
    df_novo = pd.DataFrame(registros)
    salvar_saidas(df_novo, colunas_base, paths)

    print("\n" + _color("=" * 60, "cyan"))
    print(_color("SUCESSO!", "green", bold=True))
    print(_color("=" * 60, "cyan"))
    print(f"Registros processados: {len(df_novo)}")
    print(f"Colunas: {len(df_novo.columns)}")


if __name__ == "__main__":
    main()
